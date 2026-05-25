# ============================================================
#  Marktforschungs-Dashboard v2 · Streamlit + Plotly
#  Enthält: Linie, Donut, Stacked Bar, Spider/Radar,
#           Heatmap, Bubble, Sankey, Balken, Zufriedenheit
#
#  Installation:  pip install streamlit plotly pandas
#  Starten:       streamlit run
# ============================================================

import os
from pathlib import Path
import pyreadstat
import streamlit as st
from streamlit_extras.stylable_container import stylable_container
import numpy as np
from colors import C_LIST, GFS_BLUE, CREME_FARBE
from global_css import GLOBAL
from header import header
import helpers as h
from KpiRenderer import KpiRenderer
import pickle
import plotly.express as px
from single_barchart import create_barchart
from stacked_barchart import create_stacked_bar
from density_chart import create_density_plot
import HandleMeta
from HandleMeta import load_metatable, get_group_label_and_single_labels
from environment import IN_GFS
from pie_chart import create_piechart
from big_number import big_number
from footer import footer
import plotly.express as px
from card_handler import CardHandler
from crunch_label import give_crunch_label
cards = CardHandler()
import pandas as pd
from openpyxl import load_workbook
from types import SimpleNamespace


def read_xlsx(path, sheet_name=0):
    df = pd.read_excel(path, sheet_name=sheet_name)

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else wb[sheet_name]

    cols = list(df.columns)
    dtypes = {c: str(df[c].dtype) for c in cols}

    meta = SimpleNamespace(
        # Spalten
        column_names            = cols,
        column_labels           = cols[:],              # keine echten Labels in xlsx
        column_names_to_labels  = {c: c for c in cols}, # 1:1-Mapping als Fallback
        # Dimensionen
        number_rows             = len(df),
        number_columns          = df.shape[1],
        # Typen
        original_variable_types = dtypes,
        readstat_variable_types = dtypes,
        variable_storage_width  = {c: None for c in cols},
        variable_display_width  = {c: None for c in cols},
        variable_measure        = {c: None for c in cols},
        variable_alignment      = {c: None for c in cols},
        # Labels / Codes (in Excel nicht vorhanden)
        variable_value_labels   = {},
        value_labels            = {},
        variable_to_label       = {},
        missing_ranges          = {},
        missing_user_values     = {},
        # Datei-Ebene
        file_label              = ws.title,
        file_encoding           = None,
        table_name              = ws.title,
        notes                   = [],
    )
    return df, meta

df, meta = read_xlsx("tracking.xlsx")




# ── SEITEN-KONFIGURATION ────────────────────────────────────
st.set_page_config(
    page_title="Lisas-Concentration-Dashboard",
    page_icon="face.png",
    layout="wide",
)

# ── GLOBALES CSS ─────────────────────────────────────────────
st.markdown(GLOBAL, unsafe_allow_html=True)

st.markdown(h.set_background(hex_color=CREME_FARBE[1]), unsafe_allow_html=True)

# ── HEADER ──────────────────────────────────────────────────
st.markdown(
    header(
        title="CONCENTRATION 2023 - 2025",
        subtitle="Concentration is HARD. And my psychiatrist who is no other than Milan Scheidegger asked me if the amount of concentration a day is variable. And thats a question I want to answer throughoutly, rather than skim it. So here is the full report of 2 years tracking concentration.",
        year=2026,
        window="Concentration-Dashboard",
        icon = "face.png"
    ),
    unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
#  KPI-KARTEN
# ══════════════════════════════════════════════════════════
kpi = KpiRenderer(df)
kpi.render_section_title()

k1, k2, k3, k4 = st.columns(4)

# kpi.anzahl_befragte_mit_jahr(k1)

kpi.render(k1, "Anzahl Tage", "125", "+0 % Gegenüber dem letzten Jahr", "up")
kpi.render(k2, "Methode", "SELF REPORT", "0% Veränderung", "down")
kpi.render(k3, "Bearbeitungszeit", "täglich 2 min", "0% Veränderung", "down")
kpi.render(k4, "Befragungszeitraum", "2023-2025", "im exam-period", "down")

st.divider()

# ══════════════════════════════════════════════════════════
#  GFS1_1 · MITTELWERTE NACH JAHR
# ══════════════════════════════════════════════════════════


b11, b2 = st.columns([1, 2.5])


big_number(b11, "Rest-Day", "Der Tag nach einem Rest Day (277.6 min) wird", "+90", "länger konzentriert als vorher (187.6) (n=7, p=0.07) (d = 0.65)", color="grün", add_percent="min", height=280)
    
big_number(b11, "Fluktuation", "Im letzten Semester", "31,4", "... schwingt die Lernzeit pro Tag hin und her (gegeben getrackt)", color="gelb", add_percent=True, height=280)


with b2:
    with st.container(key="karte3"):
        h.set_subtle_title(st, "Minutenanzahl", "Mittelwerte nach Break")
        import pandas as pd
        import plotly.express as px

        df_temp = df.dropna(subset=["Semester", "Minutenanzahl"])
        df_temp["Semester"] = df_temp["Semester"].astype(int)

        fig = px.line(
            df_temp,
            x="Tag-Nr.",
            y="Minutenanzahl",
            color="Semester",
            markers=True,

        )
        fig.update_layout(xaxis_title="Tag-Nr. im Semester", yaxis_title="Minuten")
        st.subheader("Lernminuten pro Tag (nach Semester)")
        st.plotly_chart(fig, use_container_width=True)
        h.set_sample_size(st, col="Minutenanzahl", meta=meta, df=df_temp)


df_agg, meta_agg = read_xlsx("tracking_aggregated.xlsx")

a_1, a_2 = st.columns([1, 1])

with a_1:
    with st.container(key="karte4"):
        h.set_subtle_title(st, "Minutenanzahl SD", "SD nach Break")
        st.subheader("""Daily SD in Minutes and amount of "Missed Days" (not tracked, either no studying, or a little and not tracked or enough stress to work without it)""")

        break_options = {
            "Semester": "Semester",
            "Woche": "Woche",
        }

        selected_break_label = st.selectbox(
            "Break auswählen:",
            list(break_options.keys())
        )
        current_break = break_options[selected_break_label]

        fig = create_barchart(df_agg, meta_agg, "sd", current_break, horizontal=False, color ="grün", height=335)
        st.plotly_chart(fig, use_container_width=True)
        fig = create_barchart(df_agg, meta_agg, "missed_days", current_break, horizontal=False, color ="gelb", height=335)
        st.plotly_chart(fig, use_container_width=True)
        h.set_sample_size(st, col="sd", meta=meta_agg, df=df_agg)
        



with a_2:
    with st.container(key="karte2"):
        h.set_subtle_title(st, "Minutenanzahl", "Mittelwerte nach Break")
        st.subheader(HandleMeta.get_column_label(meta, "Minutenanzahl"))

        break_options = {
            "Semester": "Semester",
            "Woche": "Woche",
        }

        selected_break_label = st.selectbox(
            "Break auswählen",
            list(break_options.keys())
        )
        current_break = break_options[selected_break_label]

        fig = create_barchart(df, meta, "Minutenanzahl", current_break, horizontal=False, color ="rot", height=335)
        st.plotly_chart(fig, use_container_width=True)
        h.set_sample_size(st, col="Minutenanzahl", meta=meta, df=df)

    with st.container(key="karte8"):
        h.set_subtle_title(st, "H1", "Summary and Interpretation")
        st.subheader("""what we can see here is a general increase of study hours over the weeks (within semester). This is likely due to routine forming, concentration muscles building and pressure increasing. Missed days follow a U-shape. This can be explained for the same reason, as well as in the end of the semester, the last exam might be on a monday and then we have 4 missed days. The SD is relatively stable over the semester, averaging at around 70 Minutes. Here the difference is between Semesters, as the most recent Semester has the highest SD with (82.5 mins instead of 56 mins of the first semester).As well as the missed days are higher. This generally points to a less consistent semester. Possible explanations might be a) lack of motivation over time. or b) less muscle building and finding "grouve" of work.""")

from scipy import stats


df = df.dropna(subset=["Semester"]).copy()
df["Semester"] = df["Semester"].astype(int)
df = df.sort_values(["Semester", "Tag-Nr."]).reset_index(drop=True)

# Pro Rest-Tag: Vortag + Folgetag sammeln (nur innerhalb desselben Semesters)
before, after = [], []
for sem, g in df.groupby("Semester"):
    g = g.reset_index(drop=True)
    for i in g.index[g["Rest"] == 1]:
        if i - 1 < 0 or i + 1 >= len(g):
            continue
        b = g.loc[i - 1, "Minutenanzahl"]
        a = g.loc[i + 1, "Minutenanzahl"]
        if pd.notna(b) and pd.notna(a):   # nur vollständige Paare
            before.append(b)
            after.append(a)

before = np.array(before)
after  = np.array(after)

print(f"n Paare:        {len(before)}")
print(f"Mean Vortag:    {before.mean():.1f}")
print(f"Mean Folgetag:  {after.mean():.1f}")
print(f"Mean Differenz: {(after - before).mean():+.1f}")
print(f"Overall mean:   {df['Minutenanzahl'].dropna().mean():.1f}")
print()# H1: Folgetag > Vortag (einseitig)
t, p_t = stats.ttest_rel(after, before, alternative="greater")
print(f"Paired t-test (one-sided, after > before):  t = {t:+.3f}   p = {p_t:.4f}")

w, p_w = stats.wilcoxon(after, before, alternative="greater")
print(f"Wilcoxon       (one-sided, after > before): W = {w:.1f}     p = {p_w:.4f}")


# n Paare:        7
# Mean Vortag:    187.6
# Mean Folgetag:  277.6
# Mean Differenz: +90.0
# Overall mean:   256.4

# Paired t-test (one-sided, after > before):  t = +1.728   p = 0.0674
# Wilcoxon       (one-sided, after > before): W = 21.0     p = 0.1484

diff = after - before
d_z = diff.mean() / diff.std(ddof=1)
df = len(diff) - 1
J = 1 - 3 / (4*df - 1)
g_z = J * d_z
print(f"Cohen's d_z: {d_z:.3f}   Hedges' g_z: {g_z:.3f}")